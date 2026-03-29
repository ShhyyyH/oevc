import cv2
import io
import numpy as np
from pathlib import Path
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_to_tuple


def _detect_bubbles_on_array(img_bgr, output_path=None):
    if img_bgr is None:
        return 0, None

    # 统一缩放到可控尺寸，后续阈值和面积参数更稳定。
    img = cv2.resize(img_bgr, None, fx=0.25, fy=0.25)
    img_result = img.copy()
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape

    # 1) 自动找圆盘样本区域：选取靠下的候选圆，避免固定矩形 ROI 带来的偏差。
    blur_for_circle = cv2.GaussianBlur(gray, (9, 9), 2)
    circles = cv2.HoughCircles(
        blur_for_circle,
        cv2.HOUGH_GRADIENT,
        dp=1.2,
        minDist=max(40, h // 8),
        param1=120,
        param2=35,
        minRadius=max(40, int(min(h, w) * 0.10)),
        maxRadius=int(min(h, w) * 0.48),
    )

    if circles is None:
        return 0, img_result

    circles = np.round(circles[0]).astype(int)
    sample_circle = max(circles, key=lambda c: (c[1], c[2]))
    cx, cy, r = sample_circle.tolist()

    mask = np.zeros_like(gray)
    cv2.circle(mask, (cx, cy), int(r * 0.92), 255, -1)

    # 2) 孔洞增强：CLAHE + 黑帽，突出“亮背景上的小暗孔”。
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    gray_eq = clahe.apply(gray)
    blackhat_kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (9, 9))
    bubble_enhanced = cv2.morphologyEx(gray_eq, cv2.MORPH_BLACKHAT, blackhat_kernel)

    # 3) 仅在圆盘内阈值分割，并清理孤立噪点。
    bubble_enhanced = cv2.bitwise_and(bubble_enhanced, bubble_enhanced, mask=mask)
    binary = cv2.adaptiveThreshold(
        bubble_enhanced,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        21,
        -2,
    )
    binary = cv2.bitwise_and(binary, binary, mask=mask)

    clean_kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    binary_clean = cv2.morphologyEx(binary, cv2.MORPH_OPEN, clean_kernel)

    # 4) 轮廓过滤：面积 + 圆度，剔除划痕/边缘碎片。
    contours, _ = cv2.findContours(binary_clean, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    valid_bubbles = []

    min_area = max(3, int(r * r * 0.00003))
    max_area = max(min_area + 1, int(r * r * 0.003))
    inner_r2 = (r * 0.88) * (r * 0.88)

    for c in contours:
        area = cv2.contourArea(c)
        if area < min_area or area > max_area:
            continue

        perimeter = cv2.arcLength(c, True)
        if perimeter <= 0:
            continue

        circularity = 4.0 * np.pi * area / (perimeter * perimeter)
        if circularity < 0.35:
            continue

        m = cv2.moments(c)
        if m["m00"] == 0:
            continue

        ccx = m["m10"] / m["m00"]
        ccy = m["m01"] / m["m00"]
        if (ccx - cx) * (ccx - cx) + (ccy - cy) * (ccy - cy) > inner_r2:
            continue

        valid_bubbles.append(c)

    # 5) 可视化输出。
    cv2.circle(img_result, (cx, cy), r, (255, 0, 0), 2)
    cv2.drawContours(img_result, valid_bubbles, -1, (0, 255, 0), 1)
    cv2.putText(
        img_result,
        f"bubbles: {len(valid_bubbles)}",
        (20, 40),
        cv2.FONT_HERSHEY_SIMPLEX,
        1,
        (0, 255, 255),
        2,
        cv2.LINE_AA,
    )

    if output_path:
        cv2.imwrite(str(output_path), img_result)

    return len(valid_bubbles), img_result


def process_bubble_image_robust(image_path, output_path="result.jpg"):
    img = cv2.imread(image_path)
    if img is None:
        print("图片读取失败")
        return 0, None
    return _detect_bubbles_on_array(img, output_path=output_path)


def _image_anchor_to_row_col(anchor):
    if isinstance(anchor, str):
        return coordinate_to_tuple(anchor)

    if hasattr(anchor, "_from"):
        return anchor._from.row + 1, anchor._from.col + 1

    return None, None


def _openpyxl_image_to_bgr(image_obj):
    if not hasattr(image_obj, "_data"):
        return None

    image_bytes = image_obj._data()
    pil_img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    rgb = np.array(pil_img)
    return cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)


def _process_wps_images(input_xlsx, output_xlsx, debug_dir=None):
    """WPS 格式特殊处理：从 WpsReserved_CellImgList 的 _images 提取图片"""
    in_path = Path(input_xlsx)
    out_path = Path(output_xlsx)
    wb_in = load_workbook(in_path)
    wb_out = Workbook()

    default_ws = wb_out.active
    wb_out.remove(default_ws)

    if debug_dir:
        Path(debug_dir).mkdir(parents=True, exist_ok=True)

    total_images = 0
    detected_images = 0
    results = []  # 保存所有识别结果

    # 只处理 Sheet1
    if "Sheet1" not in wb_in.sheetnames:
        print("xlsx 中未找到 Sheet1")
        return 0, 0

    ws_in = wb_in["Sheet1"]
    ws_out = wb_out.create_sheet(title="Sheet1")
    ws_map = wb_out.create_sheet(title="WPS_图片计数映射")
    ws_map.cell(row=1, column=1, value="图片序号")
    ws_map.cell(row=1, column=2, value="气泡数")

    # 从 WpsReserved_CellImgList 获取所有图片
    if "WpsReserved_CellImgList" not in wb_in.sheetnames:
        print("xlsx 中未找到 WPS 图片存储 sheet")
        return 0, 0

    wps_sheet = wb_in["WpsReserved_CellImgList"]
    images = getattr(wps_sheet, "_images", [])
    print(f"从 WPS sheet 读取 {len(images)} 个图片对象")

    # 假设图片按照某个规律排放。这里先按顺序处理，结果暂存。
    for idx, image_obj in enumerate(images, start=1):
        bgr = _openpyxl_image_to_bgr(image_obj)
        if bgr is None:
            continue

        total_images += 1
        debug_path = None
        if debug_dir:
            debug_path = Path(debug_dir) / f"img{total_images:03d}.jpg"

        count, _ = _detect_bubbles_on_array(bgr, output_path=debug_path)
        if count > 0:
            detected_images += 1

        results.append((idx, count))
        ws_map.cell(row=len(results) + 1, column=1, value=int(idx))
        ws_map.cell(row=len(results) + 1, column=2, value=int(count))

        if idx % 20 == 0:
            print(f"  已处理 {idx}/{len(images)}...")

    # Sheet1 给一个说明，避免用户看到空表困惑。
    ws_out.cell(row=1, column=1, value="该文件源于 WPS 特殊图片格式，图片位置映射不可直接恢复。")
    ws_out.cell(row=2, column=1, value="请查看 sheet: WPS_图片计数映射")

    wb_out.save(out_path)
    print(f"\n批处理完成，共处理图片: {total_images}，成功检测到气泡的图片: {detected_images}")
    print(f"结果文件已保存: {out_path}")

    # 输出识别统计（方便查看模式）
    print("\n识别结果统计（图片序号 -> 气泡数）:")
    for idx, count in results[:20]:
        print(f"  图 {idx}: {count} 个气泡")
    if len(results) > 20:
        print(f"  ... 共 {len(results)} 张图")

    return total_images, detected_images


def process_excel_embedded_images(input_xlsx, output_xlsx, debug_dir=None):
    in_path = Path(input_xlsx)
    out_path = Path(output_xlsx)
    wb_in = load_workbook(in_path)

    # 检查是否为 WPS 格式（有 WpsReserved_CellImgList）
    if "WpsReserved_CellImgList" in wb_in.sheetnames:
        print("检测到 WPS 格式文件，使用 WPS 特殊处理...")
        return _process_wps_images(input_xlsx, output_xlsx, debug_dir)

    # 标准 Excel 处理
    wb_out = Workbook()

    # 删除默认空 sheet，后面按输入文件的 sheet 重建。
    default_ws = wb_out.active
    wb_out.remove(default_ws)

    total_images = 0
    detected_images = 0

    if debug_dir:
        Path(debug_dir).mkdir(parents=True, exist_ok=True)

    for ws in wb_in.worksheets:
        ws_out = wb_out.create_sheet(title=ws.title)
        images = getattr(ws, "_images", [])

        for idx, image_obj in enumerate(images, start=1):
            row, col = _image_anchor_to_row_col(image_obj.anchor)
            if row is None or col is None:
                continue

            bgr = _openpyxl_image_to_bgr(image_obj)
            if bgr is None:
                continue

            total_images += 1
            debug_path = None
            if debug_dir:
                debug_path = Path(debug_dir) / f"{ws.title}_R{row}C{col}_{idx}.jpg"

            count, _ = _detect_bubbles_on_array(bgr, output_path=debug_path)
            if count > 0:
                detected_images += 1

            ws_out.cell(row=row, column=col, value=int(count))

    wb_out.save(out_path)
    print(f"批处理完成，共处理图片: {total_images}，成功检测到气泡的图片: {detected_images}")
    print(f"结果文件已保存: {out_path}")

    return total_images, detected_images


if __name__ == "__main__":
    # 1) 单图模式（默认）
    # 2) Excel 批量模式：python test.py --excel input.xlsx --output counts.xlsx --debug-dir debug
    import argparse

    parser = argparse.ArgumentParser(description="多孔气泡识别：单图与 Excel 批量处理")
    parser.add_argument("--image", default="buble.jpg", help="单图模式输入图片路径")
    parser.add_argument("--result", default="result.jpg", help="单图模式结果图路径")
    parser.add_argument("--excel", default=None, help="Excel 输入路径（内嵌图片）")
    parser.add_argument("--output", default="bubble_counts.xlsx", help="Excel 输出路径")
    parser.add_argument("--debug-dir", default=None, help="批量模式下保存每张图检测可视化")
    args = parser.parse_args()

    if args.excel:
        process_excel_embedded_images(args.excel, args.output, args.debug_dir)
    else:
        count, result_img = process_bubble_image_robust(args.image, args.result)
        print(f"检测到的有效气泡数: {count}")
        if result_img is not None:
            print(f"已保存可视化结果到 {args.result}")